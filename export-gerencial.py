"""
Script para exportar itens e triggers de um host do Zabbix
Gera relatório gerencial em HTML com visualização profissional
"""

import json
import argparse
import os
from datetime import datetime
from pyzabbix import ZabbixAPI
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# Configurações do Zabbix a partir de variáveis de ambiente
ZABBIX_URL = os.getenv("ZABBIX_URL", "http://localhost")
ZABBIX_USER = os.getenv("ZABBIX_USER", "Admin")
ZABBIX_PASSWORD = os.getenv("ZABBIX_PASSWORD", "zabbix")

ITEM_TYPES = {
    "0": "Zabbix agent",
    "2": "Zabbix trapper",
    "3": "Simple check",
    "5": "Zabbix internal",
    "7": "Zabbix agent (active)",
    "9": "Web item",
    "10": "External check",
    "11": "Database monitor",
    "12": "IPMI agent",
    "13": "SSH agent",
    "14": "Telnet agent",
    "15": "Calculated",
    "16": "JMX agent",
    "17": "SNMP trap",
    "18": "Dependent item",
    "19": "HTTP agent",
    "20": "SNMP agent",
    "21": "Script"
}

VALUE_TYPES = {
    "0": "Numérico (float)",
    "1": "Caractere",
    "2": "Log",
    "3": "Numérico (inteiro)",
    "4": "Texto"
}

TRIGGER_PRIORITIES = {
    "0": ("Não classificada", "#97AAB3"),
    "1": ("Informação", "#7499FF"),
    "2": ("Aviso", "#FFC859"),
    "3": ("Média", "#FFA059"),
    "4": ("Alta", "#E97659"),
    "5": ("Desastre", "#E45959")
}

STATUS_MAP = {
    "0": ("Ativo", "#4CAF50"),
    "1": ("Desabilitado", "#F44336")
}


def conectar_zabbix(url, user, password):
    try:
        zapi = ZabbixAPI(url)
        zapi.login(user, password)
        print(f"Conectado ao Zabbix API v{zapi.api_version()}")
        return zapi
    except Exception as e:
        print(f"Erro ao conectar ao Zabbix: {e}")
        raise


def obter_host_id(zapi, hostname):
    hosts = zapi.host.get(
        filter={"host": hostname},
        output=["hostid", "host", "name"]
    )
    
    if not hosts:
        raise ValueError(f"Host '{hostname}' não encontrado!")
    
    return hosts[0]


def exportar_itens(zapi, hostid):
    itens = zapi.item.get(
        hostids=hostid,
        output=[
            "itemid", "name", "key_", "type", "value_type",
            "delay", "history", "trends", "status", "description",
            "units", "params", "formula", "logtimefmt", "preprocessing"
        ],
        selectPreprocessing=["type", "params", "error_handler", "error_handler_params"],
        selectTags=["tag", "value"]
    )
    
    print(f"  - {len(itens)} itens encontrados")
    return itens


def exportar_triggers(zapi, hostid):
    triggers = zapi.trigger.get(
        hostids=hostid,
        output=[
            "triggerid", "description", "expression", "recovery_expression",
            "priority", "status", "type", "recovery_mode", "correlation_mode",
            "correlation_tag", "manual_close", "comments", "url"
        ],
        selectDependencies=["triggerid", "description"],
        selectTags=["tag", "value"],
        expandExpression=True
    )
    
    print(f"  - {len(triggers)} triggers encontradas")
    return triggers


def exportar_problemas(zapi, hostid):
    problems = zapi.problem.get(
        hostids=hostid,
        output=["eventid", "objectid", "clock", "name", "severity", "acknowledged", "r_eventid", "r_clock"],
        recent=True,
        sortfield=["eventid"],
        sortorder="DESC"
    )
    
    print(f"  - {len(problems)} problemas encontrados")
    return problems


def exportar_eventos(zapi, hostid, limit=1000):
    events = zapi.event.get(
        hostids=hostid,
        output=["eventid", "clock", "name", "severity", "acknowledged", "r_eventid"],
        source=0, 
        object=0,  
        value=1, 
        sortfield=["clock"],
        sortorder="DESC",
        limit=limit
    )
    
    print(f"  - {len(events)} eventos encontrados")
    return events


def gerar_estatisticas(itens, triggers):
    stats = {
        "itens": {
            "total": len(itens),
            "ativos": sum(1 for i in itens if i.get("status") == "0"),
            "desabilitados": sum(1 for i in itens if i.get("status") == "1"),
            "por_tipo": Counter(ITEM_TYPES.get(i.get("type"), "Desconhecido") for i in itens),
            "por_valor": Counter(VALUE_TYPES.get(i.get("value_type"), "Desconhecido") for i in itens),
            "com_preprocessing": sum(1 for i in itens if i.get("preprocessing")),
            "com_tags": sum(1 for i in itens if i.get("tags"))
        },
        "triggers": {
            "total": len(triggers),
            "ativas": sum(1 for t in triggers if t.get("status") == "0"),
            "desabilitadas": sum(1 for t in triggers if t.get("status") == "1"),
            "por_prioridade": Counter(t.get("priority", "0") for t in triggers),
            "com_dependencias": sum(1 for t in triggers if t.get("dependencies")),
            "com_tags": sum(1 for t in triggers if t.get("tags"))
        }
    }
    return stats


def gerar_html(export_data, stats, output_file):
    
    host = export_data["host"]
    info = export_data["export_info"]
    itens = export_data.get("itens", [])
    triggers = export_data.get("triggers", [])
    
    itens_sorted = sorted(itens, key=lambda x: x.get("name", "").lower())
    
    triggers_sorted = sorted(triggers, key=lambda x: (-int(x.get("priority", "0")), x.get("description", "").lower()))
    
    html = f'''<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Relatório Zabbix - {host["name"]}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
            min-height: 100vh;
            color: #e0e0e0;
            padding: 20px;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
        }}
        
        .header {{
            background: linear-gradient(135deg, #0f3460 0%, #16213e 100%);
            border-radius: 15px;
            padding: 30px;
            margin-bottom: 25px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3);
            border: 1px solid #2a4a6e;
        }}
        
        .header h1 {{
            font-size: 2.5em;
            color: #00d4ff;
            margin-bottom: 10px;
            text-shadow: 0 0 20px rgba(0, 212, 255, 0.3);
        }}
        
        .header-info {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-top: 20px;
        }}
        
        .header-info-item {{
            background: rgba(255,255,255,0.05);
            padding: 15px;
            border-radius: 10px;
            border-left: 4px solid #00d4ff;
        }}
        
        .header-info-item label {{
            font-size: 0.85em;
            color: #888;
            display: block;
            margin-bottom: 5px;
        }}
        
        .header-info-item span {{
            font-size: 1.1em;
            color: #fff;
            font-weight: 500;
        }}
        
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
            gap: 20px;
            margin-bottom: 25px;
        }}
        
        .stat-card {{
            background: linear-gradient(135deg, #1e3a5f 0%, #16213e 100%);
            border-radius: 15px;
            padding: 25px;
            box-shadow: 0 5px 20px rgba(0,0,0,0.2);
            border: 1px solid #2a4a6e;
            transition: transform 0.3s ease;
        }}
        
        .stat-card:hover {{
            transform: translateY(-5px);
        }}
        
        .stat-card h3 {{
            color: #00d4ff;
            margin-bottom: 20px;
            font-size: 1.2em;
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        
        .stat-number {{
            font-size: 3em;
            font-weight: bold;
            color: #fff;
            line-height: 1;
        }}
        
        .stat-label {{
            color: #888;
            font-size: 0.9em;
            margin-top: 5px;
        }}
        
        .stat-breakdown {{
            margin-top: 20px;
            padding-top: 15px;
            border-top: 1px solid rgba(255,255,255,0.1);
        }}
        
        .stat-row {{
            display: flex;
            justify-content: space-between;
            padding: 8px 0;
            border-bottom: 1px solid rgba(255,255,255,0.05);
        }}
        
        .stat-row:last-child {{
            border-bottom: none;
        }}
        
        .priority-badge {{
            display: inline-flex;
            align-items: center;
            gap: 5px;
            padding: 3px 10px;
            border-radius: 12px;
            font-size: 0.85em;
            font-weight: 500;
        }}
        
        .section {{
            background: linear-gradient(135deg, #1e3a5f 0%, #16213e 100%);
            border-radius: 15px;
            margin-bottom: 25px;
            box-shadow: 0 5px 20px rgba(0,0,0,0.2);
            border: 1px solid #2a4a6e;
            overflow: hidden;
        }}
        
        .section-header {{
            background: rgba(0, 212, 255, 0.1);
            padding: 20px 25px;
            border-bottom: 1px solid #2a4a6e;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        
        .section-header h2 {{
            color: #00d4ff;
            font-size: 1.4em;
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        
        .section-count {{
            background: #00d4ff;
            color: #1a1a2e;
            padding: 5px 15px;
            border-radius: 20px;
            font-weight: bold;
        }}
        
        .search-box {{
            padding: 15px 25px;
            background: rgba(0,0,0,0.2);
        }}
        
        .search-box input {{
            width: 100%;
            padding: 12px 20px;
            border: 1px solid #2a4a6e;
            border-radius: 25px;
            background: rgba(255,255,255,0.05);
            color: #fff;
            font-size: 1em;
            outline: none;
            transition: border-color 0.3s;
        }}
        
        .search-box input:focus {{
            border-color: #00d4ff;
        }}
        
        .search-box input::placeholder {{
            color: #666;
        }}
        
        .table-container {{
            overflow-x: auto;
        }}
        
        table {{
            width: 100%;
            border-collapse: collapse;
        }}
        
        th {{
            background: rgba(0, 212, 255, 0.15);
            padding: 15px;
            text-align: left;
            font-weight: 600;
            color: #00d4ff;
            font-size: 0.9em;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            position: sticky;
            top: 0;
        }}
        
        td {{
            padding: 12px 15px;
            border-bottom: 1px solid rgba(255,255,255,0.05);
            font-size: 0.95em;
        }}
        
        tr:hover {{
            background: rgba(0, 212, 255, 0.05);
        }}
        
        .status-active {{
            color: #4CAF50;
        }}
        
        .status-disabled {{
            color: #F44336;
        }}
        
        .tag {{
            display: inline-block;
            background: rgba(0, 212, 255, 0.2);
            color: #00d4ff;
            padding: 3px 8px;
            border-radius: 4px;
            font-size: 0.8em;
            margin: 2px;
        }}
        
        .item-key {{
            font-family: 'Consolas', 'Monaco', monospace;
            background: rgba(0,0,0,0.3);
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 0.85em;
            color: #ffd700;
            word-break: break-all;
        }}
        
        .trigger-expression {{
            font-family: 'Consolas', 'Monaco', monospace;
            background: rgba(0,0,0,0.3);
            padding: 8px;
            border-radius: 4px;
            font-size: 0.8em;
            color: #90EE90;
            word-break: break-all;
            max-width: 400px;
            display: block;
        }}
        
        .footer {{
            text-align: center;
            padding: 20px;
            color: #666;
            font-size: 0.9em;
        }}
        
        .chart-container {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            padding: 20px;
        }}
        
        .chart-box {{
            background: rgba(0,0,0,0.2);
            border-radius: 10px;
            padding: 20px;
        }}
        
        .chart-box h4 {{
            color: #00d4ff;
            margin-bottom: 15px;
            font-size: 1em;
        }}
        
        .bar-chart {{
            display: flex;
            flex-direction: column;
            gap: 8px;
        }}
        
        .bar-item {{
            display: flex;
            align-items: center;
            gap: 10px;
        }}
        
        .bar-label {{
            width: 150px;
            font-size: 0.85em;
            color: #aaa;
            text-overflow: ellipsis;
            overflow: hidden;
            white-space: nowrap;
        }}
        
        .bar-track {{
            flex: 1;
            height: 24px;
            background: rgba(255,255,255,0.1);
            border-radius: 12px;
            overflow: hidden;
        }}
        
        .bar-fill {{
            height: 100%;
            background: linear-gradient(90deg, #00d4ff, #0099cc);
            border-radius: 12px;
            display: flex;
            align-items: center;
            justify-content: flex-end;
            padding-right: 10px;
            font-size: 0.85em;
            font-weight: bold;
            min-width: 40px;
            transition: width 0.5s ease;
        }}
        
        .collapsible {{
            cursor: pointer;
        }}
        
        .collapsible-content {{
            max-height: 600px;
            overflow-y: auto;
        }}
        
        @media (max-width: 768px) {{
            .header h1 {{
                font-size: 1.8em;
            }}
            
            .stat-number {{
                font-size: 2em;
            }}
            
            th, td {{
                padding: 10px 8px;
                font-size: 0.85em;
            }}
        }}
        
        /* Scrollbar personalizada */
        ::-webkit-scrollbar {{
            width: 8px;
            height: 8px;
        }}
        
        ::-webkit-scrollbar-track {{
            background: rgba(0,0,0,0.2);
        }}
        
        ::-webkit-scrollbar-thumb {{
            background: #00d4ff;
            border-radius: 4px;
        }}
        
        ::-webkit-scrollbar-thumb:hover {{
            background: #0099cc;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Relatório de Monitoramento Zabbix</h1>
            <div class="header-info">
                <div class="header-info-item">
                    <label>Host</label>
                    <span>{host["name"]}</span>
                </div>
                <div class="header-info-item">
                    <label>Technical Name</label>
                    <span>{host["host"]}</span>
                </div>
                <div class="header-info-item">
                    <label>Host ID</label>
                    <span>{host["hostid"]}</span>
                </div>
                <div class="header-info-item">
                    <label>Data da Exportação</label>
                    <span>{datetime.fromisoformat(info["data_exportacao"]).strftime("%d/%m/%Y %H:%M:%S")}</span>
                </div>
            </div>
        </div>
        
        <!-- Estatísticas Principais -->
        <div class="stats-grid">
            <div class="stat-card">
                <h3>Itens de Monitoramento</h3>
                <div class="stat-number">{stats["itens"]["total"]}</div>
                <div class="stat-label">itens configurados</div>
                <div class="stat-breakdown">
                    <div class="stat-row">
                        <span>Ativos</span>
                        <span style="color: #4CAF50; font-weight: bold;">{stats["itens"]["ativos"]}</span>
                    </div>
                    <div class="stat-row">
                        <span>Desabilitados</span>
                        <span style="color: #F44336; font-weight: bold;">{stats["itens"]["desabilitados"]}</span>
                    </div>
                    <div class="stat-row">
                        <span>Com Preprocessing</span>
                        <span>{stats["itens"]["com_preprocessing"]}</span>
                    </div>
                    <div class="stat-row">
                        <span>Com Tags</span>
                        <span>{stats["itens"]["com_tags"]}</span>
                    </div>
                </div>
            </div>
            
            <div class="stat-card">
                <h3>Triggers (Alertas)</h3>
                <div class="stat-number">{stats["triggers"]["total"]}</div>
                <div class="stat-label">triggers configuradas</div>
                <div class="stat-breakdown">
                    <div class="stat-row">
                        <span>Ativas</span>
                        <span style="color: #4CAF50; font-weight: bold;">{stats["triggers"]["ativas"]}</span>
                    </div>
                    <div class="stat-row">
                        <span>Desabilitadas</span>
                        <span style="color: #F44336; font-weight: bold;">{stats["triggers"]["desabilitadas"]}</span>
                    </div>
                    <div class="stat-row">
                        <span>Com Dependências</span>
                        <span>{stats["triggers"]["com_dependencias"]}</span>
                    </div>
                    <div class="stat-row">
                        <span>Com Tags</span>
                        <span>{stats["triggers"]["com_tags"]}</span>
                    </div>
                </div>
            </div>
            
            <div class="stat-card">
                <h3>Triggers por Severidade</h3>
                <div class="stat-breakdown" style="margin-top: 0; padding-top: 0; border-top: none;">
'''
    
    for priority in ["5", "4", "3", "2", "1", "0"]:
        count = stats["triggers"]["por_prioridade"].get(priority, 0)
        name, color = TRIGGER_PRIORITIES.get(priority, ("Desconhecido", "#888"))
        percentage = (count / stats["triggers"]["total"] * 100) if stats["triggers"]["total"] > 0 else 0
        html += f'''
                    <div class="stat-row">
                        <span class="priority-badge" style="background: {color}20; color: {color};">{name}</span>
                        <span style="font-weight: bold;">{count}</span>
                    </div>'''
    
    html += '''
                </div>
            </div>
        </div>
        
        <!-- Gráficos de Distribuição -->
        <div class="section">
            <div class="section-header">
                <h2>Distribuição dos Dados</h2>
            </div>
            <div class="chart-container">
                <div class="chart-box">
                    <h4>Itens por Tipo de Coleta</h4>
                    <div class="bar-chart">
'''
    
    top_tipos = stats["itens"]["por_tipo"].most_common(8)
    max_tipo = max(count for _, count in top_tipos) if top_tipos else 1
    
    for tipo, count in top_tipos:
        percentage = (count / max_tipo * 100)
        html += f'''
                        <div class="bar-item">
                            <span class="bar-label" title="{tipo}">{tipo}</span>
                            <div class="bar-track">
                                <div class="bar-fill" style="width: {percentage}%;">{count}</div>
                            </div>
                        </div>'''
    
    html += '''
                    </div>
                </div>
                <div class="chart-box">
                    <h4>Itens por Tipo de Valor</h4>
                    <div class="bar-chart">
'''
    
    max_valor = max(stats["itens"]["por_valor"].values()) if stats["itens"]["por_valor"] else 1
    
    for valor, count in stats["itens"]["por_valor"].most_common():
        percentage = (count / max_valor * 100)
        html += f'''
                        <div class="bar-item">
                            <span class="bar-label" title="{valor}">{valor}</span>
                            <div class="bar-track">
                                <div class="bar-fill" style="width: {percentage}%; background: linear-gradient(90deg, #ffd700, #ff8c00);">{count}</div>
                            </div>
                        </div>'''
    
    html += '''
                    </div>
                </div>
            </div>
        </div>
        
        <!-- Tabela de Itens -->
        <div class="section">
            <div class="section-header">
                <h2>Itens de Monitoramento</h2>
                <span class="section-count">''' + str(len(itens)) + '''</span>
            </div>
            <div class="search-box">
                <input type="text" id="searchItens" placeholder="Pesquisar itens por nome, chave ou descrição..." onkeyup="filterTable('itensTable', this.value)">
            </div>
            <div class="table-container collapsible-content">
                <table id="itensTable">
                    <thead>
                        <tr>
                            <th>Status</th>
                            <th>Nome</th>
                            <th>Chave</th>
                            <th>Tipo</th>
                            <th>Tipo Valor</th>
                            <th>Intervalo</th>
                            <th>Histórico</th>
                            <th>Tags</th>
                        </tr>
                    </thead>
                    <tbody>
'''
    
    for item in itens_sorted:
        status_info = STATUS_MAP.get(item.get("status", "0"), ("Desconhecido", "#888"))
        tipo = ITEM_TYPES.get(item.get("type", ""), "Desconhecido")
        valor_tipo = VALUE_TYPES.get(item.get("value_type", ""), "Desconhecido")
        
        tags_html = ""
        for tag in item.get("tags", []):
            tag_value = f":{tag['value']}" if tag.get('value') else ""
            tags_html += f'<span class="tag">{tag["tag"]}{tag_value}</span>'
        
        html += f'''
                        <tr>
                            <td class="{'status-active' if item.get('status') == '0' else 'status-disabled'}">{status_info[0]}</td>
                            <td><strong>{item.get("name", "")}</strong></td>
                            <td><span class="item-key">{item.get("key_", "")}</span></td>
                            <td>{tipo}</td>
                            <td>{valor_tipo}</td>
                            <td>{item.get("delay", "-")}</td>
                            <td>{item.get("history", "-")}</td>
                            <td>{tags_html or "-"}</td>
                        </tr>'''
    
    html += '''
                    </tbody>
                </table>
            </div>
        </div>
        
        <!-- Tabela de Triggers -->
        <div class="section">
            <div class="section-header">
                <h2>Triggers (Alertas)</h2>
                <span class="section-count">''' + str(len(triggers)) + '''</span>
            </div>
            <div class="search-box">
                <input type="text" id="searchTriggers" placeholder="Pesquisar triggers por nome ou expressão..." onkeyup="filterTable('triggersTable', this.value)">
            </div>
            <div class="table-container collapsible-content">
                <table id="triggersTable">
                    <thead>
                        <tr>
                            <th>Status</th>
                            <th>Severidade</th>
                            <th>Nome</th>
                            <th>Expressão</th>
                            <th>Comentários</th>
                            <th>Tags</th>
                        </tr>
                    </thead>
                    <tbody>
'''
    
    for trigger in triggers_sorted:
        status_info = STATUS_MAP.get(trigger.get("status", "0"), ("Desconhecido", "#888"))
        priority = trigger.get("priority", "0")
        priority_info = TRIGGER_PRIORITIES.get(priority, ("Desconhecido", "#888"))
        
        tags_html = ""
        for tag in trigger.get("tags", []):
            tag_value = f":{tag['value']}" if tag.get('value') else ""
            tags_html += f'<span class="tag">{tag["tag"]}{tag_value}</span>'
        
        html += f'''
                        <tr>
                            <td class="{'status-active' if trigger.get('status') == '0' else 'status-disabled'}">{status_info[0]}</td>
                            <td>
                                <span class="priority-badge" style="background: {priority_info[1]}30; color: {priority_info[1]};">
                                    {priority_info[0]}
                                </span>
                            </td>
                            <td><strong>{trigger.get("description", "")}</strong></td>
                            <td><code class="trigger-expression">{trigger.get("expression", "")}</code></td>
                            <td>{trigger.get("comments", "") or "-"}</td>
                            <td>{tags_html or "-"}</td>
                        </tr>'''
    
    html += '''
                    </tbody>
                </table>
            </div>
        </div>
'''
    
    problems = export_data.get("problems", [])
    events = export_data.get("events", [])
    
    html += '''
        <!-- Tabela de Alertas Ativos -->
        <div class="section">
            <div class="section-header">
                <h2>Alertas Ativos</h2>
                <span class="section-count">''' + str(len(problems)) + '''</span>
            </div>
            <div class="table-container collapsible-content">
                <table id="alertasAtivosTable">
                    <thead>
                        <tr>
                            <th>Severidade</th>
                            <th>Problema</th>
                            <th>Inicio</th>
                            <th>Duração</th>
                            <th>Reconhecido</th>
                        </tr>
                    </thead>
                    <tbody>
'''
    
    if problems:
        for problem in problems:
            severity = problem.get("severity", "0")
            severity_info = TRIGGER_PRIORITIES.get(severity, ("Desconhecido", "#888"))
            
            clock = int(problem.get("clock", 0))
            inicio = datetime.fromtimestamp(clock).strftime("%d/%m/%Y %H:%M:%S") if clock else "-"
            
            if clock:
                duracao_seg = int(datetime.now().timestamp()) - clock
                dias = duracao_seg // 86400
                horas = (duracao_seg % 86400) // 3600
                minutos = (duracao_seg % 3600) // 60
                if dias > 0:
                    duracao = f"{dias}d {horas}h {minutos}m"
                elif horas > 0:
                    duracao = f"{horas}h {minutos}m"
                else:
                    duracao = f"{minutos}m"
            else:
                duracao = "-"
            
            ack = "Sim" if problem.get("acknowledged") == "1" else "Não"
            
            html += f'''
                        <tr>
                            <td>
                                <span class="priority-badge" style="background: {severity_info[1]}30; color: {severity_info[1]};">
                                    {severity_info[0]}
                                </span>
                            </td>
                            <td><strong>{problem.get("name", "")}</strong></td>
                            <td>{inicio}</td>
                            <td>{duracao}</td>
                            <td>{ack}</td>
                        </tr>'''
    else:
        html += '''
                        <tr>
                            <td colspan="5" style="text-align: center; color: #4CAF50; padding: 30px;">Nenhum alerta ativo no momento</td>
                        </tr>'''
    
    html += '''
                    </tbody>
                </table>
            </div>
        </div>
'''
    
    if events:
        alert_counter = Counter(e.get("name", "") for e in events)
        top_alerts = alert_counter.most_common(20)
        
        html += '''
        <!-- Tabela de Alertas Mais Frequentes -->
        <div class="section">
            <div class="section-header">
                <h2>Alertas Mais Frequentes (Histórico)</h2>
                <span class="section-count">Top 20</span>
            </div>
            <div class="table-container collapsible-content">
                <table id="alertasFrequentesTable">
                    <thead>
                        <tr>
                            <th>Posição</th>
                            <th>Alerta</th>
                            <th>Ocorrências</th>
                            <th>% do Total</th>
                        </tr>
                    </thead>
                    <tbody>
'''
        
        total_events = len(events)
        for pos, (alert_name, count) in enumerate(top_alerts, 1):
            percentual = (count / total_events * 100) if total_events > 0 else 0
            
            html += f'''
                        <tr>
                            <td style="text-align: center; font-weight: bold;">{pos}</td>
                            <td><strong>{alert_name}</strong></td>
                            <td style="text-align: center;">{count}</td>
                            <td>
                                <div style="display: flex; align-items: center; gap: 10px;">
                                    <div style="flex: 1; height: 20px; background: rgba(255,255,255,0.1); border-radius: 10px; overflow: hidden;">
                                        <div style="width: {percentual}%; height: 100%; background: linear-gradient(90deg, #E97659, #E45959); border-radius: 10px;"></div>
                                    </div>
                                    <span style="min-width: 50px;">{percentual:.1f}%</span>
                                </div>
                            </td>
                        </tr>'''
        
        html += '''
                    </tbody>
                </table>
            </div>
        </div>
'''
    
    html += '''
        <div class="footer">
            <p>Relatório gerado automaticamente em ''' + datetime.now().strftime("%d/%m/%Y às %H:%M:%S") + '''</p>
            <p>zbxVision by Nathan Schiavon - 2025</p>
        </div>
    </div>
    
    <script>
        function filterTable(tableId, searchText) {
            const table = document.getElementById(tableId);
            const rows = table.getElementsByTagName('tr');
            const search = searchText.toLowerCase();
            
            for (let i = 1; i < rows.length; i++) {
                const cells = rows[i].getElementsByTagName('td');
                let found = false;
                
                for (let j = 0; j < cells.length; j++) {
                    if (cells[j].textContent.toLowerCase().includes(search)) {
                        found = true;
                        break;
                    }
                }
                
                rows[i].style.display = found ? '' : 'none';
            }
        }
    </script>
</body>
</html>'''
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    
    print(f"Relatório HTML salvo em: {output_file}")


def gerar_excel(export_data, stats, output_file):
    
    host = export_data["host"]
    info = export_data["export_info"]
    itens = export_data.get("itens", [])
    triggers = export_data.get("triggers", [])
    problems = export_data.get("problems", [])
    events = export_data.get("events", [])
    
    wb = Workbook()
    
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="00D4FF")
    border = Border(
        left=Side(style='thin', color='2A4A6E'),
        right=Side(style='thin', color='2A4A6E'),
        top=Side(style='thin', color='2A4A6E'),
        bottom=Side(style='thin', color='2A4A6E')
    )
    
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    ws_resumo['A1'] = "RELATÓRIO DE MONITORAMENTO ZABBIX"
    ws_resumo['A1'].font = Font(bold=True, size=16, color="00D4FF")
    ws_resumo.merge_cells('A1:D1')
    
    row = 3
    ws_resumo[f'A{row}'] = "Host:"
    ws_resumo[f'B{row}'] = host["name"]
    ws_resumo[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws_resumo[f'A{row}'] = "Technical Name:"
    ws_resumo[f'B{row}'] = host["host"]
    ws_resumo[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws_resumo[f'A{row}'] = "Host ID:"
    ws_resumo[f'B{row}'] = host["hostid"]
    ws_resumo[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws_resumo[f'A{row}'] = "Data da Exportação:"
    ws_resumo[f'B{row}'] = datetime.fromisoformat(info["data_exportacao"]).strftime("%d/%m/%Y %H:%M:%S")
    ws_resumo[f'A{row}'].font = Font(bold=True)
    
    row += 3
    ws_resumo[f'A{row}'] = "ESTATÍSTICAS"
    ws_resumo[f'A{row}'].font = title_font
    ws_resumo.merge_cells(f'A{row}:D{row}')
    
    row += 2
    headers = ['Categoria', 'Métrica', 'Valor', 'Detalhes']
    for col, header in enumerate(headers, 1):
        cell = ws_resumo.cell(row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    stats_data = [
        ("Itens", "Total", stats["itens"]["total"], ""),
        ("", "Ativos", stats["itens"]["ativos"], ""),
        ("", "Desabilitados", stats["itens"]["desabilitados"], ""),
        ("", "Com Preprocessing", stats["itens"]["com_preprocessing"], ""),
        ("", "Com Tags", stats["itens"]["com_tags"], ""),
        ("Triggers", "Total", stats["triggers"]["total"], ""),
        ("", "Ativas", stats["triggers"]["ativas"], ""),
        ("", "Desabilitadas", stats["triggers"]["desabilitadas"], ""),
        ("", "Com Dependências", stats["triggers"]["com_dependencias"], ""),
        ("", "Com Tags", stats["triggers"]["com_tags"], ""),
        ("Alertas", "Ativos", len(problems), ""),
        ("", "Histórico (últimos eventos)", len(events), "")
    ]
    
    for data_row in stats_data:
        row += 1
        for col, value in enumerate(data_row, 1):
            cell = ws_resumo.cell(row, col, value)
            cell.border = border
            if col == 1 and value:
                cell.font = Font(bold=True)
    
    ws_resumo.column_dimensions['A'].width = 20
    ws_resumo.column_dimensions['B'].width = 30
    ws_resumo.column_dimensions['C'].width = 15
    ws_resumo.column_dimensions['D'].width = 30
    
    ws_itens = wb.create_sheet("Itens")
    
    headers = ['Status', 'Nome', 'Chave', 'Tipo', 'Tipo Valor', 'Intervalo', 'Histórico', 'Tags']
    for col, header in enumerate(headers, 1):
        cell = ws_itens.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    itens_sorted = sorted(itens, key=lambda x: x.get("name", "").lower())
    for row_idx, item in enumerate(itens_sorted, 2):
        status = "Ativo" if item.get("status") == "0" else "Desabilitado"
        tipo = ITEM_TYPES.get(item.get("type", ""), "Desconhecido")
        valor_tipo = VALUE_TYPES.get(item.get("value_type", ""), "Desconhecido")
        
        tags = ", ".join([f"{t['tag']}:{t.get('value', '')}" if t.get('value') else t['tag'] for t in item.get("tags", [])])
        
        data = [
            status,
            item.get("name", ""),
            item.get("key_", ""),
            tipo,
            valor_tipo,
            item.get("delay", "-"),
            item.get("history", "-"),
            tags or "-"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_itens.cell(row_idx, col, value)
            cell.border = border
            if col == 1:
                if status == "Ativo":
                    cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    ws_itens.column_dimensions['A'].width = 15
    ws_itens.column_dimensions['B'].width = 50
    ws_itens.column_dimensions['C'].width = 40
    ws_itens.column_dimensions['D'].width = 20
    ws_itens.column_dimensions['E'].width = 20
    ws_itens.column_dimensions['F'].width = 12
    ws_itens.column_dimensions['G'].width = 12
    ws_itens.column_dimensions['H'].width = 30
    
    ws_triggers = wb.create_sheet("Triggers")
    
    headers = ['Status', 'Severidade', 'Nome', 'Expressão', 'Comentários', 'Tags']
    for col, header in enumerate(headers, 1):
        cell = ws_triggers.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    triggers_sorted = sorted(triggers, key=lambda x: (-int(x.get("priority", "0")), x.get("description", "").lower()))
    for row_idx, trigger in enumerate(triggers_sorted, 2):
        status = "Ativo" if trigger.get("status") == "0" else "Desabilitado"
        priority = trigger.get("priority", "0")
        priority_name = TRIGGER_PRIORITIES.get(priority, ("Desconhecido", "#888"))[0]
        
        tags = ", ".join([f"{t['tag']}:{t.get('value', '')}" if t.get('value') else t['tag'] for t in trigger.get("tags", [])])
        
        data = [
            status,
            priority_name,
            trigger.get("description", ""),
            trigger.get("expression", ""),
            trigger.get("comments", "") or "-",
            tags or "-"
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_triggers.cell(row_idx, col, value)
            cell.border = border
            
            if col == 2:
                if priority == "5":  # Desastre
                    cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                elif priority == "4":  # Alta
                    cell.fill = PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid")
                elif priority == "3":  # Média
                    cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
                elif priority == "2":  # Aviso
                    cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                elif priority == "1":  # Informação
                    cell.fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    
    ws_triggers.column_dimensions['A'].width = 15
    ws_triggers.column_dimensions['B'].width = 18
    ws_triggers.column_dimensions['C'].width = 50
    ws_triggers.column_dimensions['D'].width = 60
    ws_triggers.column_dimensions['E'].width = 40
    ws_triggers.column_dimensions['F'].width = 30
    
    ws_alertas = wb.create_sheet("Alertas Ativos")
    
    headers = ['Severidade', 'Problema', 'Início', 'Duração', 'Reconhecido']
    for col, header in enumerate(headers, 1):
        cell = ws_alertas.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    if problems:
        for row_idx, problem in enumerate(problems, 2):
            severity = problem.get("severity", "0")
            severity_name = TRIGGER_PRIORITIES.get(severity, ("Desconhecido", "#888"))[0]
            
            clock = int(problem.get("clock", 0))
            inicio = datetime.fromtimestamp(clock).strftime("%d/%m/%Y %H:%M:%S") if clock else "-"
            
            if clock:
                duracao_seg = int(datetime.now().timestamp()) - clock
                dias = duracao_seg // 86400
                horas = (duracao_seg % 86400) // 3600
                minutos = (duracao_seg % 3600) // 60
                if dias > 0:
                    duracao = f"{dias}d {horas}h {minutos}m"
                elif horas > 0:
                    duracao = f"{horas}h {minutos}m"
                else:
                    duracao = f"{minutos}m"
            else:
                duracao = "-"
            
            ack = "Sim" if problem.get("acknowledged") == "1" else "Não"
            
            data = [severity_name, problem.get("name", ""), inicio, duracao, ack]
            
            for col, value in enumerate(data, 1):
                cell = ws_alertas.cell(row_idx, col, value)
                cell.border = border
                
                if col == 1:
                    if severity == "5":
                        cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                    elif severity == "4":
                        cell.fill = PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid")
                    elif severity == "3":
                        cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    else:
        ws_alertas['A2'] = "Nenhum alerta ativo no momento"
        ws_alertas['A2'].fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        ws_alertas.merge_cells('A2:E2')
        ws_alertas['A2'].alignment = Alignment(horizontal='center')
    
    ws_alertas.column_dimensions['A'].width = 18
    ws_alertas.column_dimensions['B'].width = 60
    ws_alertas.column_dimensions['C'].width = 20
    ws_alertas.column_dimensions['D'].width = 15
    ws_alertas.column_dimensions['E'].width = 15
    
    if events:
        ws_top = wb.create_sheet("Top 20 Alertas")
        
        headers = ['Posição', 'Alerta', 'Ocorrências', '% do Total']
        for col, header in enumerate(headers, 1):
            cell = ws_top.cell(1, col, header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        alert_counter = Counter(e.get("name", "") for e in events)
        top_alerts = alert_counter.most_common(20)
        total_events = len(events)
        
        for row_idx, (alert_name, count) in enumerate(top_alerts, 2):
            percentual = (count / total_events * 100) if total_events > 0 else 0
            
            data = [row_idx - 1, alert_name, count, f"{percentual:.1f}%"]
            
            for col, value in enumerate(data, 1):
                cell = ws_top.cell(row_idx, col, value)
                cell.border = border
                if col == 1:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
        
        ws_top.column_dimensions['A'].width = 12
        ws_top.column_dimensions['B'].width = 70
        ws_top.column_dimensions['C'].width = 15
        ws_top.column_dimensions['D'].width = 15
    
    wb.save(output_file)
    print(f"Relatório Excel salvo em: {output_file}")


def salvar_json(dados, filename):
    """Salva os dados em um arquivo JSON."""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(dados, f, indent=4, ensure_ascii=False)
    print(f"Dados JSON salvos em: {filename}")


def main():
    parser = argparse.ArgumentParser(
        description="Exporta itens e triggers de um host do Zabbix com relatório gerencial"
    )
    parser.add_argument(
        "hostname",
        help="Nome do host no Zabbix (technical name)"
    )
    parser.add_argument(
        "--url",
        default=ZABBIX_URL,
        help="URL do Zabbix"
    )
    parser.add_argument(
        "--user",
        default=ZABBIX_USER,
        help="Usuário do Zabbix"
    )
    parser.add_argument(
        "--password",
        default=ZABBIX_PASSWORD,
        help="Senha do Zabbix"
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Nome base do arquivo de saída (padrão: <hostname>_relatorio)"
    )
    parser.add_argument(
        "--format",
        choices=["html", "json", "excel", "both", "all"],
        default="both",
        help="Formato de saída: html, json, excel, both (html+json) ou all (padrão: both)"
    )
    
    args = parser.parse_args()
    
    print("="*60)
    print("🔌 Conectando ao Zabbix...")
    print("="*60)
    zapi = conectar_zabbix(args.url, args.user, args.password)
    
    print(f"\n Buscando host: {args.hostname}")
    host = obter_host_id(zapi, args.hostname)
    hostid = host["hostid"]
    print(f"   Host encontrado: {host['name']} (ID: {hostid})")
    
    export_data = {
        "export_info": {
            "data_exportacao": datetime.now().isoformat(),
            "zabbix_url": args.url,
            "zabbix_version": zapi.api_version()
        },
        "host": {
            "hostid": host["hostid"],
            "host": host["host"],
            "name": host["name"]
        }
    }
    
    print("\n Exportando itens...")
    export_data["itens"] = exportar_itens(zapi, hostid)
    
    print("\n Exportando triggers...")
    export_data["triggers"] = exportar_triggers(zapi, hostid)
    
    print("\n Exportando alertas ativos...")
    export_data["problems"] = exportar_problemas(zapi, hostid)
    
    print("\n Exportando histórico de eventos...")
    export_data["events"] = exportar_eventos(zapi, hostid)
    
    print("\n Gerando estatísticas...")
    stats = gerar_estatisticas(export_data["itens"], export_data["triggers"])
    
    output_base = args.output or f"{args.hostname}_relatorio"
    
    print("\n Salvando arquivos...")
    
    if args.format in ["json", "both", "all"]:
        salvar_json(export_data, f"{output_base}.json")
    
    if args.format in ["html", "both", "all"]:
        gerar_html(export_data, stats, f"{output_base}.html")
    
    if args.format in ["excel", "all"]:
        gerar_excel(export_data, stats, f"{output_base}.xlsx")
    
    # Resumo
    print("\n" + "="*60)
    print(" RESUMO DA EXPORTAÇÃO")
    print("="*60)
    print(f"     Host: {host['name']}")
    print(f"    Itens exportados: {len(export_data['itens'])}")
    print(f"      ├─ Ativos: {stats['itens']['ativos']}")
    print(f"      └─ Desabilitados: {stats['itens']['desabilitados']}")
    print(f"    Triggers exportadas: {len(export_data['triggers'])}")
    print(f"      ├─ Ativas: {stats['triggers']['ativas']}")
    print(f"      └─ Desabilitadas: {stats['triggers']['desabilitadas']}")
    print(f"    Arquivos gerados:")
    if args.format in ["json", "both", "all"]:
        print(f"      ├─ {output_base}.json")
    if args.format in ["html", "both", "all"]:
        print(f"      ├─ {output_base}.html")
    if args.format in ["excel", "all"]:
        print(f"      └─ {output_base}.xlsx")
    print("="*60)


if __name__ == "__main__":
    main()
